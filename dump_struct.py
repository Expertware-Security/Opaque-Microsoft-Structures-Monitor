import sys
import os
import subprocess
import platform
import argparse
from openpyxl import Workbook

DT_START = "<<<DUMP_START>>>"
DT_END = "<<<DUMP_END>>>"

CDB_PATH = r"C:\Program Files (x86)\Windows Kits\10\Debuggers\x64\cdb.exe"
KD_PATH = r"C:\Program Files (x86)\Windows Kits\10\Debuggers\x64\kd.exe"

USERMODE_STRUCT_FILE = "usermode_struct.txt"
USERMODE_VARIABLE_FILE = "usermode_variable.txt"
KERNELMODE_STRUCT_FILE = "kernel_mode_struct.txt"
KERNELMODE_VARIABLE_FILE = "kernel_mode_variable.txt"

DEBUG = False


def extract_output(all_text: str) -> str:
    global DEBUG
    start = all_text.rfind(DT_START)
    if start == -1:
        if DEBUG:
            print("[DEBUG] extract_output: DT_START not found")
        return ""
    start += len(DT_START)
    end = all_text.find(DT_END, start)
    if DEBUG:
        print(f"[DEBUG] extract_output: start={start}, end={end}")
    if end == -1 or end <= start:
        return ""
    return all_text[start:end]


def run_process(cmd_array):
    proc = subprocess.run(
        cmd_array,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return proc.stdout


def sanitize_sheet_name(name: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join(ch for ch in name if ch not in invalid)
    if not cleaned:
        cleaned = "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def create_unique_sheet(wb: Workbook, base_name: str):
    base_name = sanitize_sheet_name(base_name)
    if base_name not in wb.sheetnames:
        return wb.create_sheet(title=base_name)
    i = 1
    while True:
        suffix = f"_{i}"
        trimmed = base_name
        if len(trimmed) + len(suffix) > 31:
            trimmed = trimmed[: 31 - len(suffix)]
        candidate = trimmed + suffix
        if candidate not in wb.sheetnames:
            return wb.create_sheet(title=candidate)
        i += 1


def parse_dt_fields(dt_text: str):
    rows = []
    for line in dt_text.splitlines():
        stripped = line.strip()
        if "+0x" in stripped:
            parts = stripped.split()
            if len(parts) >= 2 and parts[0].startswith("+0x"):
                offset = parts[0]
                field_def = " ".join(parts[1:])
                rows.append((offset, field_def))
    return rows


def parse_address_block(text: str):
    lines = text.splitlines()
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        parts = stripped.split()
        if len(parts) >= 1:
            addr_str = parts[0].replace("`", "")
            try:
                return int(addr_str, 16)
            except ValueError:
                continue
    return None


def parse_lm_base(lm_text: str, module_name: str):
    lines = lm_text.splitlines()
    target = module_name.lower()
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        parts = stripped.split()
        if len(parts) >= 3:
            mod = parts[2].lower()
            if mod == target:
                base_str = parts[0].replace("`", "")
                try:
                    return int(base_str, 16)
                except ValueError:
                    continue
    return None


def read_list_file(path: str):
    if not os.path.isfile(path):
        return []
    entries = []
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            entries.append(stripped)
    return entries


def process_usermode_structs(wb: Workbook):
    entries = read_list_file(USERMODE_STRUCT_FILE)
    if DEBUG:
        print(f"[DEBUG] Usermode structs entries: {len(entries)}")
    for line in entries:
        parts = line.split()
        if len(parts) < 2:
            continue
        dll_path = parts[0]
        type_name = parts[1]
        if not os.path.isfile(dll_path):
            if DEBUG:
                print(f"[DEBUG] DLL not found for struct: {dll_path}")
            continue
        dll_basename = os.path.basename(dll_path)
        module_name, _ = os.path.splitext(dll_basename)
        dt_symbol = f"{module_name}!{type_name}"
        cdb_cmd = f".echo {DT_START}; dt {dt_symbol}; .echo {DT_END}; q"
        cmd_array = [
            CDB_PATH,
            "-z",
            dll_path,
            "-sx",
            "-xd",
            "-xn",
            "-xg",
            "-xi",
            "-c",
            cdb_cmd,
        ]
        output = run_process(cmd_array)
        if DEBUG:
            print(f"[DEBUG] CDB dt command for {dt_symbol}:")
            print(" ".join(cmd_array))
            print("[DEBUG] RAW OUTPUT BEGIN")
            print(output)
            print("[DEBUG] RAW OUTPUT END\n")
        dt_text = extract_output(output)
        if DEBUG:
            print(f"[DEBUG] Extracted dt_text length for {dt_symbol}: {len(dt_text)}")
        if not dt_text.strip():
            continue
        rows = parse_dt_fields(dt_text)
        if DEBUG:
            print(f"[DEBUG] Parsed {len(rows)} fields for {dt_symbol}")
        sheet_name = f"urm-{type_name}"
        ws = create_unique_sheet(wb, sheet_name)
        ws.cell(row=1, column=1, value=dt_symbol)
        ws.cell(row=2, column=1, value="Offset")
        ws.cell(row=2, column=2, value="Field")
        row_idx = 3
        if rows:
            for offset, field_def in rows:
                ws.cell(row=row_idx, column=1, value=offset)
                ws.cell(row=row_idx, column=2, value=field_def)
                row_idx += 1
        else:
            for ln in dt_text.splitlines():
                ws.cell(row=row_idx, column=1, value=ln)
                row_idx += 1


def process_kernelmode_structs(wb: Workbook):
    entries = read_list_file(KERNELMODE_STRUCT_FILE)
    if DEBUG:
        print(f"[DEBUG] Kernelmode structs entries: {len(entries)}")
    for line in entries:
        if "!" not in line:
            continue
        module_name, type_name = line.split("!", 1)
        dt_symbol = f"{module_name}!{type_name}"
        kd_cmd = f".reload; .echo {DT_START}; dt {dt_symbol}; .echo {DT_END}; q"
        cmd_array = [
            KD_PATH,
            "-kl",
            "-c",
            kd_cmd,
        ]
        output = run_process(cmd_array)
        if DEBUG:
            print(f"[DEBUG] KD dt command for {dt_symbol}:")
            print(" ".join(cmd_array))
            print("[DEBUG] RAW OUTPUT BEGIN")
            print(output)
            print("[DEBUG] RAW OUTPUT END\n")
        dt_text = extract_output(output)
        if DEBUG:
            print(f"[DEBUG] Extracted dt_text length for {dt_symbol}: {len(dt_text)}")
        if not dt_text.strip():
            continue
        rows = parse_dt_fields(dt_text)
        if DEBUG:
            print(f"[DEBUG] Parsed {len(rows)} fields for {dt_symbol}")
        sheet_name = f"krnl-{type_name}"
        ws = create_unique_sheet(wb, sheet_name)
        ws.cell(row=1, column=1, value=dt_symbol)
        ws.cell(row=2, column=1, value="Offset")
        ws.cell(row=2, column=2, value="Field")
        row_idx = 3
        if rows:
            for offset, field_def in rows:
                ws.cell(row=row_idx, column=1, value=offset)
                ws.cell(row=row_idx, column=2, value=field_def)
                row_idx += 1
        else:
            for ln in dt_text.splitlines():
                ws.cell(row=row_idx, column=1, value=ln)
                row_idx += 1


def process_usermode_variables(wb: Workbook):
    entries = read_list_file(USERMODE_VARIABLE_FILE)
    if DEBUG:
        print(f"[DEBUG] Usermode variables entries: {len(entries)}")
    if not entries:
        return
    ws = create_unique_sheet(wb, "usermode_variables")
    ws.cell(row=1, column=1, value="ModulePath")
    ws.cell(row=1, column=2, value="Module")
    ws.cell(row=1, column=3, value="Symbol")
    ws.cell(row=1, column=4, value="OffsetHex")
    row_idx = 2
    for line in entries:
        parts = line.split()
        if len(parts) < 2:
            continue
        dll_path = parts[0]
        symbol_name = parts[1]
        if not os.path.isfile(dll_path):
            if DEBUG:
                print(f"[DEBUG] DLL not found for variable: {dll_path}")
            continue
        dll_basename = os.path.basename(dll_path)
        module_name, _ = os.path.splitext(dll_basename)
        cdb_cmd = f".echo {DT_START}; x {module_name}!{symbol_name}; lm m {module_name}; .echo {DT_END}; q"
        cmd_array = [
            CDB_PATH,
            "-z",
            dll_path,
            "-sx",
            "-xd",
            "-xn",
            "-xg",
            "-xi",
            "-c",
            cdb_cmd,
        ]
        output = run_process(cmd_array)
        if DEBUG:
            print(f"[DEBUG] CDB x/lm command for {module_name}!{symbol_name}:")
            print(" ".join(cmd_array))
            print("[DEBUG] RAW OUTPUT BEGIN")
            print(output)
            print("[DEBUG] RAW OUTPUT END\n")
        inner = extract_output(output)
        if DEBUG:
            print(f"[DEBUG] Extracted inner length for {module_name}!{symbol_name}: {len(inner)}")
        if not inner.strip():
            continue
        lines = inner.splitlines()
        x_lines = []
        lm_lines = []
        lm_section = False
        for l in lines:
            low = l.lower()
            if low.startswith("start") and "module name" in low:
                lm_section = True
                lm_lines.append(l)
                continue
            if lm_section:
                lm_lines.append(l)
            else:
                x_lines.append(l)
        x_text = "\n".join(x_lines)
        lm_text = "\n".join(lm_lines)
        addr = parse_address_block(x_text)
        base = parse_lm_base(lm_text, module_name)
        if DEBUG:
            print(f"[DEBUG] Parsed addr/base for {module_name}!{symbol_name}: addr={addr}, base={base}")
        if addr is None or base is None:
            continue
        offset = addr - base
        offset_hex = f"0x{offset:x}"
        ws.cell(row=row_idx, column=1, value=dll_path)
        ws.cell(row=row_idx, column=2, value=module_name)
        ws.cell(row=row_idx, column=3, value=symbol_name)
        ws.cell(row=row_idx, column=4, value=offset_hex)
        row_idx += 1


def process_kernelmode_variables(wb: Workbook):
    entries = read_list_file(KERNELMODE_VARIABLE_FILE)
    if DEBUG:
        print(f"[DEBUG] Kernelmode variables entries: {len(entries)}")
    if not entries:
        return
    ws = create_unique_sheet(wb, "kernelmode_variables")
    ws.cell(row=1, column=1, value="Module")
    ws.cell(row=1, column=2, value="Symbol")
    ws.cell(row=1, column=3, value="OffsetHex")
    row_idx = 2
    for line in entries:
        if "!" not in line:
            continue
        module_name, symbol_name = line.split("!", 1)
        kd_cmd = f".reload; .echo {DT_START}; x {module_name}!{symbol_name}; lm m {module_name}; .echo {DT_END}; q"
        cmd_array = [
            KD_PATH,
            "-kl",
            "-c",
            kd_cmd,
        ]
        output = run_process(cmd_array)
        if DEBUG:
            print(f"[DEBUG] KD x/lm command for {module_name}!{symbol_name}:")
            print(" ".join(cmd_array))
            print("[DEBUG] RAW OUTPUT BEGIN")
            print(output)
            print("[DEBUG] RAW OUTPUT END\n")
        inner = extract_output(output)
        if DEBUG:
            print(f"[DEBUG] Extracted inner length for {module_name}!{symbol_name}: {len(inner)}")
        if not inner.strip():
            continue
        lines = inner.splitlines()
        x_lines = []
        lm_lines = []
        lm_section = False
        for l in lines:
            low = l.lower()
            if low.startswith("start") and "module name" in low:
                lm_section = True
                lm_lines.append(l)
                continue
            if lm_section:
                lm_lines.append(l)
            else:
                x_lines.append(l)
        x_text = "\n".join(x_lines)
        lm_text = "\n".join(lm_lines)
        addr = parse_address_block(x_text)
        base = parse_lm_base(lm_text, module_name)
        if DEBUG:
            print(f"[DEBUG] Parsed addr/base for {module_name}!{symbol_name}: addr={addr}, base={base}")
        if addr is None or base is None:
            continue
        offset = addr - base
        offset_hex = f"0x{offset:x}"
        ws.cell(row=row_idx, column=1, value=module_name)
        ws.cell(row=row_idx, column=2, value=symbol_name)
        ws.cell(row=row_idx, column=3, value=offset_hex)
        row_idx += 1


def main():
    global DEBUG
    parser = argparse.ArgumentParser()
    parser.add_argument("-debug", action="store_true", help="Enable debug output")
    args = parser.parse_args()
    DEBUG = args.debug

    build = platform.win32_ver()[1]
    if not build:
        try:
            build = str(sys.getwindowsversion().build)
        except Exception:
            build = "windows_build"
    filename = f"{build}.xlsx"

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    process_usermode_structs(wb)
    print("[+] Processed usermode structs")

    process_kernelmode_structs(wb)
    print("[+] Processed kernelmode structs")

    process_usermode_variables(wb)
    print("[+] Processed usermode variables")

    process_kernelmode_variables(wb)
    print("[+] Processed kernelmode variables")

    if DEBUG:
        print(f"[DEBUG] Final Excel sheets: {wb.sheetnames}")

    wb.save(filename)
    print(f"[+] Saved Excel file: {filename}")


if __name__ == "__main__":
    main()
